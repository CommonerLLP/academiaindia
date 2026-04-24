"""Build the institutions registry.

IMPORTANT: Every URL in this file is a scaffolding guess from the author's
memory. NONE have been verified against the live institutional website at
build time. The `coverage_status` column is set to 'Unverified' across the
board. Run `verify_registry.py` to probe each URL, update HTTP status, and
promote verified rows to 'Stub' or 'Active' based on whether a parser exists.

Why this pattern: relying on my training-data URLs directly would launch a
job-tracker with systematically broken links. Flagging every row as
unverified and running a probe step inverts the risk — the registry is
born mistrusted and earns its reliability.
"""

from __future__ import annotations

from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@dataclass
class Institution:
    id: str
    name: str
    short_name: str
    type: str  # IIT | IIM | IISER | IISc | CentralUniversity | NIT | IIIT | AIIMS | Other
    state: str
    city: str
    established: Optional[int]
    statute_basis: str
    career_page_url_guess: str  # to be verified
    ad_format_guess: str = "Unknown"  # HTML | PDFOnly | Samarth | Mixed | Unknown
    parser: str = "generic"
    coverage_status: str = "Unverified"
    notes: str = ""


# ---- Institutions ----
# Convention: id is slug-form; name is canonical long; URLs are all to be verified.
# I intentionally did NOT fabricate a URL where I was uncertain; such rows get
# career_page_url_guess = "" and coverage_status = "Unverified (no URL guess)".

IITS = [
    Institution("iit-kharagpur", "Indian Institute of Technology Kharagpur", "IIT Kharagpur", "IIT", "West Bengal", "Kharagpur", 1951, "IITAct1961", "https://iitkgp.ac.in/jobs"),
    Institution("iit-bombay", "Indian Institute of Technology Bombay", "IIT Bombay", "IIT", "Maharashtra", "Mumbai", 1958, "IITAct1961", "https://www.iitb.ac.in/en/careers"),
    Institution("iit-madras", "Indian Institute of Technology Madras", "IIT Madras", "IIT", "Tamil Nadu", "Chennai", 1959, "IITAct1961", "https://www.iitm.ac.in/jobs"),
    Institution("iit-kanpur", "Indian Institute of Technology Kanpur", "IIT Kanpur", "IIT", "Uttar Pradesh", "Kanpur", 1959, "IITAct1961", "https://www.iitk.ac.in/jobs"),
    Institution("iit-delhi", "Indian Institute of Technology Delhi", "IIT Delhi", "IIT", "Delhi", "New Delhi", 1961, "IITAct1961", "https://home.iitd.ac.in/jobs.php", parser="iit_delhi"),
    Institution("iit-guwahati", "Indian Institute of Technology Guwahati", "IIT Guwahati", "IIT", "Assam", "Guwahati", 1994, "IITAct1961", "https://www.iitg.ac.in/recruitment"),
    Institution("iit-roorkee", "Indian Institute of Technology Roorkee", "IIT Roorkee", "IIT", "Uttarakhand", "Roorkee", 1847, "IITAct1961", "https://iitr.ac.in/Jobs"),
    Institution("iit-bhubaneswar", "Indian Institute of Technology Bhubaneswar", "IIT Bhubaneswar", "IIT", "Odisha", "Bhubaneswar", 2008, "IITAct1961", "https://www.iitbbs.ac.in/jobs.php"),
    Institution("iit-gandhinagar", "Indian Institute of Technology Gandhinagar", "IIT Gandhinagar", "IIT", "Gujarat", "Gandhinagar", 2008, "IITAct1961", "https://iitgn.ac.in/jobs"),
    Institution("iit-hyderabad", "Indian Institute of Technology Hyderabad", "IIT Hyderabad", "IIT", "Telangana", "Sangareddy", 2008, "IITAct1961", "https://iith.ac.in/jobs/"),
    Institution("iit-jodhpur", "Indian Institute of Technology Jodhpur", "IIT Jodhpur", "IIT", "Rajasthan", "Jodhpur", 2008, "IITAct1961", "https://www.iitj.ac.in/jobs"),
    Institution("iit-patna", "Indian Institute of Technology Patna", "IIT Patna", "IIT", "Bihar", "Patna", 2008, "IITAct1961", "https://www.iitp.ac.in/jobs"),
    Institution("iit-ropar", "Indian Institute of Technology Ropar", "IIT Ropar", "IIT", "Punjab", "Rupnagar", 2008, "IITAct1961", "https://www.iitrpr.ac.in/jobs"),
    Institution("iit-indore", "Indian Institute of Technology Indore", "IIT Indore", "IIT", "Madhya Pradesh", "Indore", 2009, "IITAct1961", "https://www.iiti.ac.in/page/job-opportunities"),
    Institution("iit-mandi", "Indian Institute of Technology Mandi", "IIT Mandi", "IIT", "Himachal Pradesh", "Mandi", 2009, "IITAct1961", "https://www.iitmandi.ac.in/jobs"),
    Institution("iit-bhilai", "Indian Institute of Technology Bhilai", "IIT Bhilai", "IIT", "Chhattisgarh", "Durg", 2016, "IITAct1961", "https://www.iitbhilai.ac.in/jobs"),
    Institution("iit-goa", "Indian Institute of Technology Goa", "IIT Goa", "IIT", "Goa", "Ponda", 2016, "IITAct1961", "https://www.iitgoa.ac.in/jobs"),
    Institution("iit-jammu", "Indian Institute of Technology Jammu", "IIT Jammu", "IIT", "Jammu and Kashmir", "Jammu", 2016, "IITAct1961", "https://www.iitjammu.ac.in/jobs"),
    Institution("iit-dharwad", "Indian Institute of Technology Dharwad", "IIT Dharwad", "IIT", "Karnataka", "Dharwad", 2016, "IITAct1961", "https://www.iitdh.ac.in/jobs"),
    Institution("iit-palakkad", "Indian Institute of Technology Palakkad", "IIT Palakkad", "IIT", "Kerala", "Palakkad", 2015, "IITAct1961", "https://iitpkd.ac.in/jobs"),
    Institution("iit-tirupati", "Indian Institute of Technology Tirupati", "IIT Tirupati", "IIT", "Andhra Pradesh", "Tirupati", 2015, "IITAct1961", "https://iittp.ac.in/jobs"),
    Institution("iit-ism-dhanbad", "Indian Institute of Technology (ISM) Dhanbad", "IIT-ISM Dhanbad", "IIT", "Jharkhand", "Dhanbad", 1926, "IITAct1961", "https://www.iitism.ac.in/jobs"),
    Institution("iit-varanasi-bhu", "Indian Institute of Technology (BHU) Varanasi", "IIT-BHU Varanasi", "IIT", "Uttar Pradesh", "Varanasi", 1919, "IITAct1961", "https://iitbhu.ac.in/recruitments"),
]

IIMS = [
    Institution("iim-calcutta", "Indian Institute of Management Calcutta", "IIM Calcutta", "IIM", "West Bengal", "Kolkata", 1961, "IIMAct2017", "https://www.iimcal.ac.in/jobs"),
    Institution("iim-ahmedabad", "Indian Institute of Management Ahmedabad", "IIM Ahmedabad", "IIM", "Gujarat", "Ahmedabad", 1961, "IIMAct2017", "https://www.iima.ac.in/careers"),
    Institution("iim-bangalore", "Indian Institute of Management Bangalore", "IIM Bangalore", "IIM", "Karnataka", "Bengaluru", 1973, "IIMAct2017", "https://www.iimb.ac.in/careers"),
    Institution("iim-lucknow", "Indian Institute of Management Lucknow", "IIM Lucknow", "IIM", "Uttar Pradesh", "Lucknow", 1984, "IIMAct2017", "https://www.iiml.ac.in/careers"),
    Institution("iim-kozhikode", "Indian Institute of Management Kozhikode", "IIM Kozhikode", "IIM", "Kerala", "Kozhikode", 1996, "IIMAct2017", "https://www.iimk.ac.in/careers"),
    Institution("iim-indore", "Indian Institute of Management Indore", "IIM Indore", "IIM", "Madhya Pradesh", "Indore", 1996, "IIMAct2017", "https://www.iimidr.ac.in/careers"),
    Institution("iim-shillong", "Indian Institute of Management Shillong", "IIM Shillong", "IIM", "Meghalaya", "Shillong", 2007, "IIMAct2017", "https://www.iimshillong.ac.in/careers"),
    Institution("iim-rohtak", "Indian Institute of Management Rohtak", "IIM Rohtak", "IIM", "Haryana", "Rohtak", 2010, "IIMAct2017", "https://www.iimrohtak.ac.in/careers"),
    Institution("iim-ranchi", "Indian Institute of Management Ranchi", "IIM Ranchi", "IIM", "Jharkhand", "Ranchi", 2010, "IIMAct2017", "https://www.iimranchi.ac.in/careers"),
    Institution("iim-raipur", "Indian Institute of Management Raipur", "IIM Raipur", "IIM", "Chhattisgarh", "Raipur", 2010, "IIMAct2017", "https://www.iimraipur.ac.in/careers"),
    Institution("iim-trichy", "Indian Institute of Management Tiruchirappalli", "IIM Trichy", "IIM", "Tamil Nadu", "Tiruchirappalli", 2011, "IIMAct2017", "https://www.iimtrichy.ac.in/careers"),
    Institution("iim-kashipur", "Indian Institute of Management Kashipur", "IIM Kashipur", "IIM", "Uttarakhand", "Kashipur", 2011, "IIMAct2017", "https://www.iimkashipur.ac.in/careers"),
    Institution("iim-udaipur", "Indian Institute of Management Udaipur", "IIM Udaipur", "IIM", "Rajasthan", "Udaipur", 2011, "IIMAct2017", "https://www.iimu.ac.in/careers"),
    Institution("iim-amritsar", "Indian Institute of Management Amritsar", "IIM Amritsar", "IIM", "Punjab", "Amritsar", 2015, "IIMAct2017", "https://www.iimamritsar.ac.in/careers"),
    Institution("iim-bodhgaya", "Indian Institute of Management Bodh Gaya", "IIM Bodh Gaya", "IIM", "Bihar", "Bodh Gaya", 2015, "IIMAct2017", "https://www.iimbg.ac.in/careers"),
    Institution("iim-nagpur", "Indian Institute of Management Nagpur", "IIM Nagpur", "IIM", "Maharashtra", "Nagpur", 2015, "IIMAct2017", "https://www.iimnagpur.ac.in/careers"),
    Institution("iim-sambalpur", "Indian Institute of Management Sambalpur", "IIM Sambalpur", "IIM", "Odisha", "Sambalpur", 2015, "IIMAct2017", "https://www.iimsambalpur.ac.in/careers"),
    Institution("iim-sirmaur", "Indian Institute of Management Sirmaur", "IIM Sirmaur", "IIM", "Himachal Pradesh", "Sirmaur", 2015, "IIMAct2017", "https://www.iimsirmaur.ac.in/careers"),
    Institution("iim-visakhapatnam", "Indian Institute of Management Visakhapatnam", "IIM Visakhapatnam", "IIM", "Andhra Pradesh", "Visakhapatnam", 2015, "IIMAct2017", "https://www.iimv.ac.in/careers"),
    Institution("iim-jammu", "Indian Institute of Management Jammu", "IIM Jammu", "IIM", "Jammu and Kashmir", "Jammu", 2016, "IIMAct2017", "https://www.iimj.ac.in/careers"),
    Institution("iim-mumbai", "Indian Institute of Management Mumbai (formerly NITIE)", "IIM Mumbai", "IIM", "Maharashtra", "Mumbai", 1963, "IIMAct2017", "https://www.iimmumbai.ac.in/careers"),
]

IISERS_IISC = [
    Institution("iisc-bangalore", "Indian Institute of Science Bangalore", "IISc Bangalore", "IISc", "Karnataka", "Bengaluru", 1909, "Individual", "https://iisc.ac.in/careers/"),
    Institution("iiser-pune", "Indian Institute of Science Education and Research Pune", "IISER Pune", "IISER", "Maharashtra", "Pune", 2006, "Individual", "https://www.iiserpune.ac.in/jobs"),
    Institution("iiser-kolkata", "Indian Institute of Science Education and Research Kolkata", "IISER Kolkata", "IISER", "West Bengal", "Mohanpur", 2006, "Individual", "https://www.iiserkol.ac.in/jobs"),
    Institution("iiser-mohali", "Indian Institute of Science Education and Research Mohali", "IISER Mohali", "IISER", "Punjab", "Mohali", 2007, "Individual", "https://www.iisermohali.ac.in/jobs"),
    Institution("iiser-bhopal", "Indian Institute of Science Education and Research Bhopal", "IISER Bhopal", "IISER", "Madhya Pradesh", "Bhopal", 2008, "Individual", "https://www.iiserb.ac.in/jobs"),
    Institution("iiser-thiruvananthapuram", "Indian Institute of Science Education and Research Thiruvananthapuram", "IISER TVM", "IISER", "Kerala", "Thiruvananthapuram", 2008, "Individual", "https://www.iisertvm.ac.in/jobs"),
    Institution("iiser-tirupati", "Indian Institute of Science Education and Research Tirupati", "IISER Tirupati", "IISER", "Andhra Pradesh", "Tirupati", 2015, "Individual", "https://www.iisertirupati.ac.in/jobs"),
    Institution("iiser-berhampur", "Indian Institute of Science Education and Research Berhampur", "IISER Berhampur", "IISER", "Odisha", "Berhampur", 2016, "Individual", "https://www.iiserbpr.ac.in/jobs"),
]

# Central Universities — set from CUs listed on MoE site as of training cutoff.
# List completeness is approximate; several newer CUs (e.g. in Andhra Pradesh,
# Telangana post-bifurcation) may be missing or misnamed. Flag notes below.
CENTRAL_UNIVERSITIES = [
    Institution("jnu", "Jawaharlal Nehru University", "JNU", "CentralUniversity", "Delhi", "New Delhi", 1969, "Individual", "https://www.jnu.ac.in/recruitment", parser="jnu"),
    Institution("du", "University of Delhi", "DU", "CentralUniversity", "Delhi", "New Delhi", 1922, "Individual", "https://www.du.ac.in/index.php?page=vacancies"),
    Institution("jmi", "Jamia Millia Islamia", "JMI", "CentralUniversity", "Delhi", "New Delhi", 1920, "Individual", "https://www.jmi.ac.in/jobs"),
    Institution("bhu", "Banaras Hindu University", "BHU", "CentralUniversity", "Uttar Pradesh", "Varanasi", 1916, "Individual", "https://bhu.ac.in/jobs"),
    Institution("amu", "Aligarh Muslim University", "AMU", "CentralUniversity", "Uttar Pradesh", "Aligarh", 1920, "Individual", "https://www.amu.ac.in/career"),
    Institution("visva-bharati", "Visva-Bharati", "Visva-Bharati", "CentralUniversity", "West Bengal", "Santiniketan", 1921, "Individual", "https://www.visvabharati.ac.in/jobs"),
    Institution("nehu", "North-Eastern Hill University", "NEHU", "CentralUniversity", "Meghalaya", "Shillong", 1973, "Individual", "https://nehu.ac.in/jobs"),
    Institution("hcu", "University of Hyderabad", "HCU", "CentralUniversity", "Telangana", "Hyderabad", 1974, "Individual", "https://www.uohyd.ac.in/index.php/notices/employment"),
    Institution("pondicherry-u", "Pondicherry University", "Pondicherry U", "CentralUniversity", "Puducherry", "Puducherry", 1985, "Individual", "https://www.pondiuni.edu.in/jobs"),
    Institution("assam-u", "Assam University", "Assam U", "CentralUniversity", "Assam", "Silchar", 1994, "Individual", "https://www.aus.ac.in/jobs"),
    Institution("tezpur-u", "Tezpur University", "Tezpur U", "CentralUniversity", "Assam", "Tezpur", 1994, "Individual", "https://www.tezu.ernet.in/jobs"),
    Institution("nagaland-u", "Nagaland University", "Nagaland U", "CentralUniversity", "Nagaland", "Lumami", 1994, "Individual", "https://nagalanduniversity.ac.in/jobs"),
    Institution("mizoram-u", "Mizoram University", "Mizoram U", "CentralUniversity", "Mizoram", "Aizawl", 2001, "Individual", "https://www.mzu.edu.in/jobs"),
    Institution("manipur-u", "Manipur University", "Manipur U", "CentralUniversity", "Manipur", "Imphal", 1980, "Individual", "https://manipuruniv.ac.in/jobs"),
    Institution("tripura-u", "Tripura University", "Tripura U", "CentralUniversity", "Tripura", "Agartala", 1987, "Individual", "https://www.tripurauniv.ac.in/jobs"),
    Institution("sikkim-u", "Sikkim University", "Sikkim U", "CentralUniversity", "Sikkim", "Gangtok", 2007, "CentralUnivsAct2009", "https://www.cus.ac.in/jobs"),
    Institution("hnbgu", "Hemvati Nandan Bahuguna Garhwal University", "HNBGU", "CentralUniversity", "Uttarakhand", "Srinagar Garhwal", 1973, "Individual", "https://hnbgu.ac.in/jobs"),
    Institution("mahatma-gandhi-antarrashtriya-hindi-u", "Mahatma Gandhi Antarrashtriya Hindi Vishwavidyalaya", "MGAHV", "CentralUniversity", "Maharashtra", "Wardha", 1997, "Individual", "https://www.hindivishwa.org/jobs"),
    Institution("igntu", "Indira Gandhi National Tribal University", "IGNTU", "CentralUniversity", "Madhya Pradesh", "Amarkantak", 2007, "CentralUnivsAct2009", "https://www.igntu.ac.in/jobs"),
    Institution("cu-bihar", "Central University of Bihar / Mahatma Gandhi Central University", "MGCU/CUB", "CentralUniversity", "Bihar", "Motihari", 2009, "CentralUnivsAct2009", "https://mgcub.ac.in/jobs"),
    Institution("cu-gujarat", "Central University of Gujarat", "CU Gujarat", "CentralUniversity", "Gujarat", "Gandhinagar", 2009, "CentralUnivsAct2009", "https://cug.ac.in/jobs"),
    Institution("cu-haryana", "Central University of Haryana", "CU Haryana", "CentralUniversity", "Haryana", "Mahendragarh", 2009, "CentralUnivsAct2009", "https://cuh.ac.in/jobs"),
    Institution("cu-himachal", "Central University of Himachal Pradesh", "CU HP", "CentralUniversity", "Himachal Pradesh", "Dharamshala", 2009, "CentralUnivsAct2009", "https://cuhimachal.ac.in/jobs"),
    Institution("cu-jammu", "Central University of Jammu", "CU Jammu", "CentralUniversity", "Jammu and Kashmir", "Samba", 2009, "CentralUnivsAct2009", "https://cujammu.ac.in/jobs"),
    Institution("cu-kashmir", "Central University of Kashmir", "CU Kashmir", "CentralUniversity", "Jammu and Kashmir", "Ganderbal", 2009, "CentralUnivsAct2009", "https://cukashmir.ac.in/jobs"),
    Institution("cu-jharkhand", "Central University of Jharkhand", "CU Jharkhand", "CentralUniversity", "Jharkhand", "Ranchi", 2009, "CentralUnivsAct2009", "https://cuj.ac.in/jobs"),
    Institution("cu-karnataka", "Central University of Karnataka", "CU Karnataka", "CentralUniversity", "Karnataka", "Kalaburagi", 2009, "CentralUnivsAct2009", "https://cuk.ac.in/jobs"),
    Institution("cu-kerala", "Central University of Kerala", "CU Kerala", "CentralUniversity", "Kerala", "Kasaragod", 2009, "CentralUnivsAct2009", "https://cukerala.ac.in/jobs"),
    Institution("cu-odisha", "Central University of Odisha", "CU Odisha", "CentralUniversity", "Odisha", "Koraput", 2009, "CentralUnivsAct2009", "https://cuo.ac.in/jobs"),
    Institution("cu-punjab", "Central University of Punjab", "CU Punjab", "CentralUniversity", "Punjab", "Bathinda", 2009, "CentralUnivsAct2009", "https://cup.edu.in/jobs"),
    Institution("cu-rajasthan", "Central University of Rajasthan", "CU Rajasthan", "CentralUniversity", "Rajasthan", "Ajmer", 2009, "CentralUnivsAct2009", "https://curaj.ac.in/jobs"),
    Institution("cu-tamilnadu", "Central University of Tamil Nadu", "CU TN", "CentralUniversity", "Tamil Nadu", "Thiruvarur", 2009, "CentralUnivsAct2009", "https://cutn.ac.in/jobs"),
    Institution("cu-ap", "Central University of Andhra Pradesh", "CU AP", "CentralUniversity", "Andhra Pradesh", "Ananthapuramu", 2018, "CentralUnivsAct2009", "https://cuap.ac.in/jobs", notes="Post-bifurcation CU; URL speculative."),
    Institution("cu-telangana", "University of Hyderabad-adjacent; CU Telangana (?)", "CU Telangana", "CentralUniversity", "Telangana", "N/A", None, "CentralUnivsAct2009", "", notes="Uncertain whether a separate CU Telangana was notified distinct from UoH. Verify against MoE list."),
    Institution("rgu-arunachal", "Rajiv Gandhi University", "RGU", "CentralUniversity", "Arunachal Pradesh", "Itanagar", 1984, "Individual", "https://rgu.ac.in/jobs"),
    Institution("ignou", "Indira Gandhi National Open University", "IGNOU", "CentralUniversity", "Delhi", "New Delhi", 1985, "Individual", "https://www.ignou.ac.in/ignou/aboutignou/recruitment"),
    Institution("epch", "English and Foreign Languages University", "EFLU", "CentralUniversity", "Telangana", "Hyderabad", 1958, "Individual", "https://efluniversity.ac.in/jobs"),
    Institution("mbbu", "Maulana Azad National Urdu University", "MANUU", "CentralUniversity", "Telangana", "Hyderabad", 1998, "Individual", "https://manuu.edu.in/jobs"),
    Institution("rgnau", "Rajiv Gandhi National Aviation University", "RGNAU", "CentralUniversity", "Uttar Pradesh", "Amethi", 2013, "Individual", "https://rgnau.ac.in/jobs", notes="Aviation-specific CU; faculty recruitment sparse."),
    Institution("nalanda-u", "Nalanda University", "Nalanda U", "CentralUniversity", "Bihar", "Rajgir", 2014, "Individual", "https://nalandauniv.edu.in/jobs", notes="Distinct Nalanda University Act 2010; not strictly CU under CU Act 2009."),
    Institution("cust-bilaspur", "Guru Ghasidas Vishwavidyalaya", "GGV", "CentralUniversity", "Chhattisgarh", "Bilaspur", 1983, "Individual", "https://ggu.ac.in/jobs"),
    Institution("harisingh-gour-u", "Dr. Harisingh Gour Vishwavidyalaya", "DHSGSU", "CentralUniversity", "Madhya Pradesh", "Sagar", 1946, "Individual", "https://dhsgsu.edu.in/jobs"),
    Institution("bbau-lucknow", "Babasaheb Bhimrao Ambedkar University", "BBAU", "CentralUniversity", "Uttar Pradesh", "Lucknow", 1996, "Individual", "https://www.bbau.ac.in/jobs"),
    Institution("rgv-u-allahabad", "University of Allahabad", "Allahabad U", "CentralUniversity", "Uttar Pradesh", "Prayagraj", 1887, "Individual", "https://www.allduniv.ac.in/jobs"),
    Institution("hp-u-shimla", "Himachal Pradesh Central University (rename pending)", "HPCU", "CentralUniversity", "Himachal Pradesh", "Shimla", None, "CentralUnivsAct2009", "", notes="Distinct from CU HP Dharamshala? Verify existence."),
    Institution("cug-gandhinagar-alt", "DUPLICATE-CHECK CU Gujarat", "CU Gujarat dup", "CentralUniversity", "Gujarat", "Gandhinagar", 2009, "CentralUnivsAct2009", "", notes="Probable duplicate of cu-gujarat. Dedupe during verification."),
]

NITS = [
    Institution("nit-trichy", "National Institute of Technology Tiruchirappalli", "NIT Trichy", "NIT", "Tamil Nadu", "Tiruchirappalli", 1964, "NITAct2007", "https://www.nitt.edu/home/employmentnews/"),
    Institution("nit-warangal", "National Institute of Technology Warangal", "NIT Warangal", "NIT", "Telangana", "Warangal", 1959, "NITAct2007", "https://www.nitw.ac.in/jobs"),
    Institution("nit-surathkal", "National Institute of Technology Karnataka Surathkal", "NITK Surathkal", "NIT", "Karnataka", "Mangaluru", 1960, "NITAct2007", "https://www.nitk.ac.in/jobs"),
    Institution("nit-calicut", "National Institute of Technology Calicut", "NIT Calicut", "NIT", "Kerala", "Kozhikode", 1961, "NITAct2007", "https://nitc.ac.in/jobs"),
    Institution("nit-rourkela", "National Institute of Technology Rourkela", "NIT Rourkela", "NIT", "Odisha", "Rourkela", 1961, "NITAct2007", "https://nitrkl.ac.in/Academic/JobOpenings.aspx"),
    Institution("nit-durgapur", "National Institute of Technology Durgapur", "NIT Durgapur", "NIT", "West Bengal", "Durgapur", 1960, "NITAct2007", "https://nitdgp.ac.in/recruitment"),
    Institution("nit-allahabad", "Motilal Nehru National Institute of Technology Allahabad", "MNNIT Allahabad", "NIT", "Uttar Pradesh", "Prayagraj", 1961, "NITAct2007", "https://www.mnnit.ac.in/jobs"),
    Institution("nit-jaipur", "Malaviya National Institute of Technology Jaipur", "MNIT Jaipur", "NIT", "Rajasthan", "Jaipur", 1963, "NITAct2007", "https://mnit.ac.in/jobs"),
    Institution("nit-bhopal", "Maulana Azad National Institute of Technology Bhopal", "MANIT Bhopal", "NIT", "Madhya Pradesh", "Bhopal", 1960, "NITAct2007", "https://www.manit.ac.in/jobs"),
    Institution("nit-nagpur", "Visvesvaraya National Institute of Technology Nagpur", "VNIT Nagpur", "NIT", "Maharashtra", "Nagpur", 1960, "NITAct2007", "https://vnit.ac.in/jobs"),
    Institution("nit-kurukshetra", "National Institute of Technology Kurukshetra", "NIT Kurukshetra", "NIT", "Haryana", "Kurukshetra", 1963, "NITAct2007", "https://nitkkr.ac.in/jobs"),
    Institution("nit-jamshedpur", "National Institute of Technology Jamshedpur", "NIT Jamshedpur", "NIT", "Jharkhand", "Jamshedpur", 1960, "NITAct2007", "https://www.nitjsr.ac.in/jobs"),
    Institution("nit-jalandhar", "Dr. B R Ambedkar National Institute of Technology Jalandhar", "NIT Jalandhar", "NIT", "Punjab", "Jalandhar", 1987, "NITAct2007", "https://www.nitj.ac.in/jobs"),
    Institution("nit-hamirpur", "National Institute of Technology Hamirpur", "NIT Hamirpur", "NIT", "Himachal Pradesh", "Hamirpur", 1986, "NITAct2007", "https://nith.ac.in/jobs"),
    Institution("nit-silchar", "National Institute of Technology Silchar", "NIT Silchar", "NIT", "Assam", "Silchar", 1967, "NITAct2007", "https://www.nits.ac.in/jobs"),
    Institution("nit-srinagar", "National Institute of Technology Srinagar", "NIT Srinagar", "NIT", "Jammu and Kashmir", "Srinagar", 1960, "NITAct2007", "https://nitsri.ac.in/jobs"),
    Institution("nit-patna", "National Institute of Technology Patna", "NIT Patna", "NIT", "Bihar", "Patna", 2004, "NITAct2007", "https://www.nitp.ac.in/jobs"),
    Institution("nit-raipur", "National Institute of Technology Raipur", "NIT Raipur", "NIT", "Chhattisgarh", "Raipur", 1956, "NITAct2007", "https://nitrr.ac.in/jobs"),
    Institution("nit-agartala", "National Institute of Technology Agartala", "NIT Agartala", "NIT", "Tripura", "Agartala", 1965, "NITAct2007", "https://www.nita.ac.in/jobs"),
    Institution("nit-delhi", "National Institute of Technology Delhi", "NIT Delhi", "NIT", "Delhi", "New Delhi", 2010, "NITAct2007", "https://nitdelhi.ac.in/jobs"),
    Institution("nit-arunachal", "National Institute of Technology Arunachal Pradesh", "NIT Arunachal", "NIT", "Arunachal Pradesh", "Jote", 2010, "NITAct2007", "https://www.nitap.ac.in/jobs"),
    Institution("nit-goa", "National Institute of Technology Goa", "NIT Goa", "NIT", "Goa", "Farmagudi", 2010, "NITAct2007", "https://www.nitgoa.ac.in/jobs"),
    Institution("nit-manipur", "National Institute of Technology Manipur", "NIT Manipur", "NIT", "Manipur", "Imphal", 2010, "NITAct2007", "https://www.nitmanipur.ac.in/jobs"),
    Institution("nit-meghalaya", "National Institute of Technology Meghalaya", "NIT Meghalaya", "NIT", "Meghalaya", "Shillong", 2010, "NITAct2007", "https://www.nitm.ac.in/jobs"),
    Institution("nit-mizoram", "National Institute of Technology Mizoram", "NIT Mizoram", "NIT", "Mizoram", "Aizawl", 2010, "NITAct2007", "https://www.nitmz.ac.in/jobs"),
    Institution("nit-nagaland", "National Institute of Technology Nagaland", "NIT Nagaland", "NIT", "Nagaland", "Dimapur", 2010, "NITAct2007", "https://www.nitnagaland.ac.in/jobs"),
    Institution("nit-puducherry", "National Institute of Technology Puducherry", "NIT Puducherry", "NIT", "Puducherry", "Karaikal", 2010, "NITAct2007", "https://www.nitpy.ac.in/jobs"),
    Institution("nit-sikkim", "National Institute of Technology Sikkim", "NIT Sikkim", "NIT", "Sikkim", "Ravangla", 2010, "NITAct2007", "https://www.nitsikkim.ac.in/jobs"),
    Institution("nit-uttarakhand", "National Institute of Technology Uttarakhand", "NIT Uttarakhand", "NIT", "Uttarakhand", "Srinagar Garhwal", 2009, "NITAct2007", "https://nituk.ac.in/jobs"),
    Institution("nit-andhra", "National Institute of Technology Andhra Pradesh", "NIT AP", "NIT", "Andhra Pradesh", "Tadepalligudem", 2015, "NITAct2007", "https://nitandhra.ac.in/jobs"),
    Institution("iiest-shibpur", "Indian Institute of Engineering Science and Technology Shibpur", "IIEST Shibpur", "NIT", "West Bengal", "Howrah", 1856, "NITAct2007", "https://www.iiests.ac.in/jobs", notes="Administratively an NIT-category institute."),
]

IIITS = [
    Institution("iiit-allahabad", "Indian Institute of Information Technology Allahabad", "IIIT Allahabad", "IIIT", "Uttar Pradesh", "Prayagraj", 1999, "IIIAct2014", "https://www.iiita.ac.in/jobs"),
    Institution("iiit-bangalore", "International Institute of Information Technology Bangalore", "IIIT Bangalore", "IIIT", "Karnataka", "Bengaluru", 1999, "Individual", "https://www.iiitb.ac.in/careers", notes="Not under IIIT Act 2014; society-mode."),
    Institution("iiit-hyderabad", "International Institute of Information Technology Hyderabad", "IIIT Hyderabad", "IIIT", "Telangana", "Hyderabad", 1998, "Individual", "https://www.iiit.ac.in/careers/", notes="Deemed; not under IIIT Act 2014."),
    Institution("iiit-delhi", "Indraprastha Institute of Information Technology Delhi", "IIIT Delhi", "IIIT", "Delhi", "New Delhi", 2008, "Individual", "https://www.iiitd.ac.in/careers", notes="State legislation; not under IIIT Act 2014."),
    Institution("iiit-gwalior", "ABV-Indian Institute of Information Technology and Management Gwalior", "ABV-IIITM Gwalior", "IIIT", "Madhya Pradesh", "Gwalior", 1997, "IIIAct2014", "https://www.iiitm.ac.in/jobs"),
    Institution("iiit-jabalpur", "Indian Institute of Information Technology Design and Manufacturing Jabalpur", "IIITDM Jabalpur", "IIIT", "Madhya Pradesh", "Jabalpur", 2005, "IIIAct2014", "https://iiitdmj.ac.in/jobs"),
    Institution("iiit-kancheepuram", "Indian Institute of Information Technology Design and Manufacturing Kancheepuram", "IIITDM Kancheepuram", "IIIT", "Tamil Nadu", "Chennai", 2007, "IIIAct2014", "https://www.iiitdm.ac.in/jobs"),
    Institution("iiit-lucknow", "Indian Institute of Information Technology Lucknow", "IIIT Lucknow", "IIIT", "Uttar Pradesh", "Lucknow", 2015, "IIIAct2014", "https://iiitl.ac.in/jobs"),
    Institution("iiit-bhagalpur", "Indian Institute of Information Technology Bhagalpur", "IIIT Bhagalpur", "IIIT", "Bihar", "Bhagalpur", 2017, "IIIAct2014", "https://iiitbh.ac.in/jobs"),
    Institution("iiit-bhopal", "Indian Institute of Information Technology Bhopal", "IIIT Bhopal", "IIIT", "Madhya Pradesh", "Bhopal", 2017, "IIIAct2014", "https://iiitbhopal.ac.in/jobs"),
    Institution("iiit-dharwad", "Indian Institute of Information Technology Dharwad", "IIIT Dharwad", "IIIT", "Karnataka", "Dharwad", 2015, "IIIAct2014", "https://iiitdwd.ac.in/jobs"),
    Institution("iiit-guwahati", "Indian Institute of Information Technology Guwahati", "IIIT Guwahati", "IIIT", "Assam", "Guwahati", 2013, "IIIAct2014", "https://www.iiitg.ac.in/jobs"),
    Institution("iiit-kalyani", "Indian Institute of Information Technology Kalyani", "IIIT Kalyani", "IIIT", "West Bengal", "Kalyani", 2014, "IIIAct2014", "https://iiitkalyani.ac.in/jobs"),
    Institution("iiit-kottayam", "Indian Institute of Information Technology Kottayam", "IIIT Kottayam", "IIIT", "Kerala", "Kottayam", 2015, "IIIAct2014", "https://iiitkottayam.ac.in/jobs"),
    Institution("iiit-kurnool", "Indian Institute of Information Technology Kurnool", "IIIT Kurnool", "IIIT", "Andhra Pradesh", "Kurnool", 2015, "IIIAct2014", "https://iiitk.ac.in/jobs"),
    Institution("iiit-manipur", "Indian Institute of Information Technology Manipur", "IIIT Manipur", "IIIT", "Manipur", "Imphal", 2015, "IIIAct2014", "https://iiitmanipur.ac.in/jobs"),
    Institution("iiit-nagpur", "Indian Institute of Information Technology Nagpur", "IIIT Nagpur", "IIIT", "Maharashtra", "Nagpur", 2016, "IIIAct2014", "https://iiitn.ac.in/jobs"),
    Institution("iiit-pune", "Indian Institute of Information Technology Pune", "IIIT Pune", "IIIT", "Maharashtra", "Pune", 2016, "IIIAct2014", "https://iiitp.ac.in/jobs"),
    Institution("iiit-ranchi", "Indian Institute of Information Technology Ranchi", "IIIT Ranchi", "IIIT", "Jharkhand", "Ranchi", 2016, "IIIAct2014", "https://iiitranchi.ac.in/jobs"),
    Institution("iiit-sonepat", "Indian Institute of Information Technology Sonepat", "IIIT Sonepat", "IIIT", "Haryana", "Sonepat", 2014, "IIIAct2014", "https://iiitsonepat.ac.in/jobs"),
    Institution("iiit-sricity", "Indian Institute of Information Technology Sri City", "IIIT Sri City", "IIIT", "Andhra Pradesh", "Sri City", 2013, "IIIAct2014", "https://www.iiits.ac.in/jobs"),
    Institution("iiit-surat", "Indian Institute of Information Technology Surat", "IIIT Surat", "IIIT", "Gujarat", "Surat", 2017, "IIIAct2014", "https://iiitsurat.ac.in/jobs"),
    Institution("iiit-tiruchirappalli", "Indian Institute of Information Technology Tiruchirappalli", "IIIT Trichy", "IIIT", "Tamil Nadu", "Tiruchirappalli", 2013, "IIIAct2014", "https://iiitt.ac.in/jobs"),
    Institution("iiit-una", "Indian Institute of Information Technology Una", "IIIT Una", "IIIT", "Himachal Pradesh", "Una", 2014, "IIIAct2014", "https://iiitu.ac.in/jobs"),
    Institution("iiit-vadodara", "Indian Institute of Information Technology Vadodara", "IIIT Vadodara", "IIIT", "Gujarat", "Gandhinagar", 2013, "IIIAct2014", "https://iiitvadodara.ac.in/jobs"),
    Institution("iiit-kota", "Indian Institute of Information Technology Kota", "IIIT Kota", "IIIT", "Rajasthan", "Kota", 2013, "IIIAct2014", "https://iiitkota.ac.in/jobs"),
]

AIIMSS = [
    Institution("aiims-delhi", "All India Institute of Medical Sciences Delhi", "AIIMS Delhi", "AIIMS", "Delhi", "New Delhi", 1956, "Individual", "https://www.aiims.edu/en/recruitment.html"),
    Institution("aiims-bhopal", "All India Institute of Medical Sciences Bhopal", "AIIMS Bhopal", "AIIMS", "Madhya Pradesh", "Bhopal", 2012, "Individual", "https://aiimsbhopal.edu.in/recruitment"),
    Institution("aiims-bhubaneswar", "All India Institute of Medical Sciences Bhubaneswar", "AIIMS Bhubaneswar", "AIIMS", "Odisha", "Bhubaneswar", 2012, "Individual", "https://aiimsbhubaneswar.nic.in/recruitment"),
    Institution("aiims-jodhpur", "All India Institute of Medical Sciences Jodhpur", "AIIMS Jodhpur", "AIIMS", "Rajasthan", "Jodhpur", 2012, "Individual", "https://www.aiimsjodhpur.edu.in/recruitment"),
    Institution("aiims-patna", "All India Institute of Medical Sciences Patna", "AIIMS Patna", "AIIMS", "Bihar", "Patna", 2012, "Individual", "https://aiimspatna.edu.in/recruitment"),
    Institution("aiims-raipur", "All India Institute of Medical Sciences Raipur", "AIIMS Raipur", "AIIMS", "Chhattisgarh", "Raipur", 2012, "Individual", "https://aiimsraipur.edu.in/recruitment"),
    Institution("aiims-rishikesh", "All India Institute of Medical Sciences Rishikesh", "AIIMS Rishikesh", "AIIMS", "Uttarakhand", "Rishikesh", 2012, "Individual", "https://aiimsrishikesh.edu.in/recruitment"),
    Institution("aiims-rae-bareli", "All India Institute of Medical Sciences Rae Bareli", "AIIMS Rae Bareli", "AIIMS", "Uttar Pradesh", "Rae Bareli", 2013, "Individual", "https://aiimsrbl.edu.in/recruitment"),
    Institution("aiims-nagpur", "All India Institute of Medical Sciences Nagpur", "AIIMS Nagpur", "AIIMS", "Maharashtra", "Nagpur", 2018, "Individual", "https://aiimsnagpur.edu.in/recruitment"),
    Institution("aiims-mangalagiri", "All India Institute of Medical Sciences Mangalagiri", "AIIMS Mangalagiri", "AIIMS", "Andhra Pradesh", "Mangalagiri", 2018, "Individual", "https://aiimsmangalagiri.edu.in/recruitment"),
    Institution("aiims-kalyani", "All India Institute of Medical Sciences Kalyani", "AIIMS Kalyani", "AIIMS", "West Bengal", "Kalyani", 2019, "Individual", "https://aiimskalyani.edu.in/recruitment"),
    Institution("aiims-gorakhpur", "All India Institute of Medical Sciences Gorakhpur", "AIIMS Gorakhpur", "AIIMS", "Uttar Pradesh", "Gorakhpur", 2018, "Individual", "https://aiimsgorakhpur.edu.in/recruitment"),
    Institution("aiims-bathinda", "All India Institute of Medical Sciences Bathinda", "AIIMS Bathinda", "AIIMS", "Punjab", "Bathinda", 2019, "Individual", "https://aiimsbathinda.edu.in/recruitment"),
    Institution("aiims-bibinagar", "All India Institute of Medical Sciences Bibinagar", "AIIMS Bibinagar", "AIIMS", "Telangana", "Bibinagar", 2018, "Individual", "https://aiimsbibinagar.edu.in/recruitment"),
    Institution("aiims-bilaspur", "All India Institute of Medical Sciences Bilaspur", "AIIMS Bilaspur", "AIIMS", "Himachal Pradesh", "Bilaspur", 2020, "Individual", "https://aiimsbilaspur.edu.in/recruitment"),
    Institution("aiims-deoghar", "All India Institute of Medical Sciences Deoghar", "AIIMS Deoghar", "AIIMS", "Jharkhand", "Deoghar", 2019, "Individual", "https://aiimsdeoghar.edu.in/recruitment"),
    Institution("aiims-guwahati", "All India Institute of Medical Sciences Guwahati", "AIIMS Guwahati", "AIIMS", "Assam", "Guwahati", 2017, "Individual", "https://aiimsguwahati.ac.in/recruitment"),
    Institution("aiims-madurai", "All India Institute of Medical Sciences Madurai", "AIIMS Madurai", "AIIMS", "Tamil Nadu", "Madurai", 2018, "Individual", "", notes="Under construction; recruitment page may not exist."),
    Institution("aiims-rajkot", "All India Institute of Medical Sciences Rajkot", "AIIMS Rajkot", "AIIMS", "Gujarat", "Rajkot", 2020, "Individual", "https://aiimsrajkot.edu.in/recruitment"),
    Institution("aiims-vijaypur", "All India Institute of Medical Sciences Vijaypur (Jammu)", "AIIMS Vijaypur", "AIIMS", "Jammu and Kashmir", "Samba", 2019, "Individual", "https://aiimsjammu.edu.in/recruitment"),
    Institution("aiims-awantipora", "All India Institute of Medical Sciences Awantipora", "AIIMS Awantipora", "AIIMS", "Jammu and Kashmir", "Pulwama", 2019, "Individual", "", notes="Site may not yet exist."),
]


def all_institutions() -> list[Institution]:
    items: list[Institution] = []
    items += IITS
    items += IIMS
    items += IISERS_IISC
    items += CENTRAL_UNIVERSITIES
    items += NITS
    items += IIITS
    items += AIIMSS
    return items


def write_xlsx(path: Path) -> dict:
    institutions = all_institutions()

    wb = Workbook()

    # --- Sheet 1: Registry ---
    ws = wb.active
    ws.title = "Registry"

    headers = [
        "id", "name", "short_name", "type", "state", "city",
        "established", "statute_basis", "career_page_url_guess",
        "ad_format_guess", "parser", "coverage_status", "notes",
    ]

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, inst in enumerate(institutions, start=2):
        record = asdict(inst)
        for col_idx, header in enumerate(headers, start=1):
            value = record.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header == "coverage_status" and value == "Unverified":
                cell.fill = PatternFill("solid", start_color="FFF2CC")

    # Column widths
    widths = {
        "id": 28, "name": 55, "short_name": 22, "type": 18, "state": 20, "city": 22,
        "established": 12, "statute_basis": 18, "career_page_url_guess": 55,
        "ad_format_guess": 16, "parser": 22, "coverage_status": 20, "notes": 50,
    }
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 16)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --- Sheet 2: Provenance ---
    pv = wb.create_sheet("Provenance")
    pv["A1"] = "Provenance and trust notes"
    pv["A1"].font = Font(name="Arial", bold=True, size=14)
    notes = [
        "",
        "This registry was seeded from the author's training-data memory.",
        "NONE of the career_page_url_guess entries have been HTTP-verified at build time.",
        "Every row is flagged coverage_status='Unverified'.",
        "",
        "Run scraper/verify_registry.py to:",
        "  1. Probe each URL (HEAD → GET fallback)",
        "  2. Record HTTP status, final redirect URL, page title",
        "  3. Flag rows where the page clearly does not exist or is not a recruitment page",
        "  4. Promote rows that resolve successfully to coverage_status='Stub'",
        "  5. Rows that pass parser smoke-tests can be promoted to 'Active'",
        "",
        "Counts by type (approximate — verify against MoE/UGC canonical list):",
        f"  IIT: {len(IITS)}",
        f"  IIM: {len(IIMS)}",
        f"  IISER/IISc: {len(IISERS_IISC)}",
        f"  Central Universities (partial): {len(CENTRAL_UNIVERSITIES)}",
        f"  NIT: {len(NITS)}",
        f"  IIIT: {len(IIITS)}",
        f"  AIIMS: {len(AIIMSS)}",
        f"  TOTAL: {len(all_institutions())}",
        "",
        "Known gaps:",
        "  - Central Universities list incomplete; MoE gazettes ~56 CUs but this file has fewer.",
        "  - Several post-2019 institutions may be missing or misnamed.",
        "  - Some IIITs under PPP-mode are subject to different recruitment practices.",
        "  - AIIMS list reflects operationally-functional institutions; announced-but-not-started ones excluded.",
        "",
        "Citations for institutional lists:",
        "  University Grants Commission. 2024. 'List of Central Universities.' UGC, New Delhi.",
        "  Ministry of Education, Government of India. 2024. 'Institutions of National Importance.' MoE, New Delhi.",
        "  All India Survey on Higher Education (AISHE) 2021-22. DHE, MoE. New Delhi.",
    ]
    for i, line in enumerate(notes, start=2):
        c = pv.cell(row=i, column=1, value=line)
        c.font = Font(name="Arial")
    pv.column_dimensions["A"].width = 100

    wb.save(path)

    return {
        "total": len(all_institutions()),
        "by_type": {
            "IIT": len(IITS), "IIM": len(IIMS), "IISER/IISc": len(IISERS_IISC),
            "CentralUniversity": len(CENTRAL_UNIVERSITIES), "NIT": len(NITS),
            "IIIT": len(IIITS), "AIIMS": len(AIIMSS),
        },
        "path": str(path),
    }


def write_json(path: Path) -> None:
    import json
    records = [asdict(inst) for inst in all_institutions()]
    path.write_text(json.dumps(records, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    base = Path(__file__).resolve().parent.parent
    xlsx_path = base / "data" / "institutions_registry.xlsx"
    json_path = base / "data" / "institutions_registry.json"
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    stats = write_xlsx(xlsx_path)
    write_json(json_path)
    import json
    print(json.dumps(stats, indent=2))
